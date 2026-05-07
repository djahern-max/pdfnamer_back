"""
routers/bills_report.py  –  Bills to Enter Report
--------------------------------------------------
Returns all confirmed PDF namings that are not cc_receipt or check_receipt,
formatted for easy QuickBooks bill entry.

Routes:
  GET /api/bills-report/pending      All bills needing QB entry (JSON)
  GET /api/bills-report/export.xlsx  Download as Excel workbook
"""

import io
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, date
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from database import get_db
from auth import require_tenant
from models.tenant import PdfNaming, Tenant

router = APIRouter(prefix="/api/bills-report", tags=["Bills Report"])

_SKIP_TYPES = {"cc_receipt", "check_receipt"}


def _fmt_date(raw: str | None) -> str:
    """Convert MMDDYYYY → MM/DD/YYYY for QuickBooks."""
    if not raw or not re.match(r"^\d{8}$", raw):
        return raw or ""
    return f"{raw[0:2]}/{raw[2:4]}/{raw[4:8]}"


class BillEntry(BaseModel):
    vendor: str | None
    bill_date: str | None  # formatted MM/DD/YYYY
    bill_no: str | None
    amount: str | None
    doc_type: str | None
    confirmed_name: str


class BillsReportResponse(BaseModel):
    bills: list[BillEntry]
    total: int
    total_amount: float


class BillItem(BaseModel):
    confirmed_name: str
    vendor: str | None
    invoice_number: str | None
    amount: str | None
    doc_date: str | None
    doc_type: str | None


def _to_item(r: PdfNaming) -> BillItem:
    return BillItem(
        confirmed_name=r.confirmed_name,
        vendor=r.vendor,
        invoice_number=r.invoice_number,
        amount=r.amount,
        doc_date=_fmt_date(r.doc_date),
        doc_type=r.doc_type,
    )


@router.get("/pending", response_model=list[BillItem])
async def pending_bills(
    tenant: Tenant = Depends(require_tenant),
    db: AsyncSession = Depends(get_db),
    since: date | None = Query(
        default=None,
        description="Filter to bills confirmed on or after this date (YYYY-MM-DD)",
    ),
):
    stmt = (
        select(PdfNaming)
        .where(
            PdfNaming.tenant_id == tenant.id,
            PdfNaming.confirmed_name.isnot(None),
            ~PdfNaming.doc_type.in_(_SKIP_TYPES),
        )
        .order_by(PdfNaming.created_at.desc())
    )
    if since:
        stmt = stmt.where(
            PdfNaming.created_at >= datetime.combine(since, datetime.min.time())
        )

    result = await db.execute(stmt)
    rows = result.scalars().all()
    return [_to_item(r) for r in rows]


@router.get("/export.xlsx")
async def export_bills_excel(
    tenant: Tenant = Depends(require_tenant),
    db: AsyncSession = Depends(get_db),
):
    """Download all pending bills as a formatted Excel workbook."""
    data = await get_pending_bills(tenant, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills to Enter"

    # ── Styles ────────────────────────────────────────────────────────────
    blue = "2563EB"
    light_blue = "DBEAFE"
    header_bg = "1E3A5F"
    row_alt = "F0F4FF"
    green = "16A34A"
    border_col = "D1D5DB"

    thin = Side(style="thin", color=border_col)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_style(cell):
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=header_bg)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = cell_border

    def title_style(cell):
        cell.font = Font(bold=True, color="FFFFFF", size=13)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def vendor_style(cell):
        cell.font = Font(bold=True, color=blue, size=10)
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = cell_border

    def total_style(cell, is_amount=False):
        cell.font = Font(bold=True, size=10, color="111827")
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        cell.alignment = Alignment(
            horizontal="right" if is_amount else "left", vertical="center"
        )
        cell.border = cell_border

    # ── Title row ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"Bills to Enter — Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    )
    title_style(title_cell)
    ws.row_dimensions[1].height = 28

    # ── Sub-title row ─────────────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"{data.total} bills  ·  Total: ${data.total_amount:,.2f}"
    sub.font = Font(italic=True, color="6B7280", size=9)
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Column headers ────────────────────────────────────────────────────
    headers = ["Vendor", "Bill Date", "Bill No.", "Amount", "Type", "File Name"]
    ws.row_dimensions[3].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        hdr_style(c)

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = [32, 14, 20, 14, 14, 46]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows grouped by vendor ───────────────────────────────────────
    # Build vendor groups
    groups: dict[str, list] = {}
    for b in data.bills:
        v = b.vendor or "Unknown"
        groups.setdefault(v, []).append(b)

    row = 4
    for vendor in sorted(groups.keys()):
        bills = groups[vendor]
        v_total = sum(float(b.amount or 0) for b in bills)

        # Vendor header row
        ws.merge_cells(f"A{row}:F{row}")
        vc = ws.cell(
            row=row,
            column=1,
            value=f"  {vendor}  —  {len(bills)} bill{'s' if len(bills)!=1 else ''}  ·  ${v_total:,.2f}",
        )
        vendor_style(vc)
        ws.row_dimensions[row].height = 18
        row += 1

        # Bill rows
        for i, b in enumerate(bills):
            fill = PatternFill("solid", fgColor=row_alt if i % 2 == 0 else "FFFFFF")
            amt = float(b.amount or 0)

            values = [
                b.vendor or "",
                b.bill_date or "",
                b.bill_no or "",
                amt,
                b.doc_type or "",
                b.confirmed_name + ".pdf",
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fill
                c.border = cell_border
                c.font = Font(size=10)

                if col == 4:  # Amount column
                    c.number_format = '"$"#,##0.00'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif col == 2:  # Date column
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 3:  # Bill No.
                    c.font = Font(size=10, bold=True)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

            ws.row_dimensions[row].height = 16
            row += 1

    # ── Grand total row ───────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    lbl = ws.cell(row=row, column=1, value=f"TOTAL  ({data.total} bills)")
    total_style(lbl)
    ws.cell(row=row, column=2).border = cell_border
    ws.cell(row=row, column=3).border = cell_border

    amt_cell = ws.cell(row=row, column=4, value=data.total_amount)
    total_style(amt_cell, is_amount=True)
    amt_cell.number_format = '"$"#,##0.00'
    amt_cell.font = Font(bold=True, size=11, color=green)

    for col in [5, 6]:
        c = ws.cell(row=row, column=col)
        total_style(c)

    ws.row_dimensions[row].height = 20

    # ── Freeze panes below headers ────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Stream response ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bills_to_enter_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
