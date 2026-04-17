"""
routers/qb_checker.py  –  QuickBooks Bill Checker (multi-tenant)
-----------------------------------------------------------------
Compares a tenant's confirmed PDF namings against a QuickBooks
Unpaid Bills Report export to identify which invoices still need
to be manually entered in QuickBooks.

Rules:
  - cc_receipt and check_receipt doc types are SKIPPED entirely
    (credit cards auto-import; cash/check receipts don't need entry)
  - All other confirmed doc types (invoice, statement, contract,
    estimate, other) are checked against the QB export

Matching strategy:
  1. Primary:  normalize(vendor) + normalize(invoice_number)
  2. Fallback: normalize(vendor) + normalize(amount)
     (for bills with no invoice number on either side)

Routes:
  POST /api/qb-checker/compare   Upload QB Excel → get needs_entry / already_entered
"""

import io
import re

import pandas as pd
from fastapi import APIRouter, File, HTTPException, UploadFile, Depends
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc

from database import get_db
from auth import require_tenant
from models.tenant import PdfNaming, Tenant

router = APIRouter(prefix="/api/qb-checker", tags=["QB Checker"])

# Doc types that never need QB entry
_SKIP_DOC_TYPES = {"cc_receipt", "check_receipt"}


# ─── Schemas ──────────────────────────────────────────────────────────────────


class BillItem(BaseModel):
    confirmed_name: str
    vendor: str | None
    invoice_number: str | None
    amount: str | None
    doc_date: str | None
    doc_type: str | None


class CompareResponse(BaseModel):
    needs_entry: list[BillItem]
    already_entered: list[BillItem]
    skipped_receipts: int
    qb_bills_parsed: int


# ─── QB Excel Parser ──────────────────────────────────────────────────────────


def _normalize(value: str | None) -> str:
    """Lowercase, strip whitespace, remove common invoice prefixes."""
    if not value:
        return ""
    s = str(value).lower().strip()
    # Strip leading "INV", "INV-", "#", "inv " etc.
    s = re.sub(r"^(inv[-#\s]?|#\s*)", "", s)
    # Strip leading zeros
    s = s.lstrip("0") or "0"
    return s


def _normalize_vendor(value: str | None) -> str:
    """Lowercase, collapse spaces, strip punctuation for fuzzy vendor match."""
    if not value:
        return ""
    s = str(value).lower().strip()
    # Remove common suffixes that vary between QB and PDF
    s = re.sub(r"\b(llc|inc|corp|co|ltd|pllc|ent|company|supply)\b\.?", "", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_qb_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse a QuickBooks Unpaid Bills Report Excel export.

    Structure:
      - Header rows at top (report title, company name, date range)
      - Column header row: NaN | Date | Transaction type | Num | Due date | Past due | Amount | Open balance
      - Vendor header rows: vendor name in col 0, rest NaN
      - Detail rows: col 0 is NaN, data in cols 1-7
      - Total rows: col 0 starts with "Total for "
    """
    df = pd.read_excel(io.BytesIO(file_bytes), header=None)

    bills = []
    current_vendor = None

    for _, row in df.iterrows():
        col0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        col1 = row.iloc[1] if pd.notna(row.iloc[1]) else None  # Date
        col2 = (
            str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
        )  # Transaction type
        col3 = (
            str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
        )  # Num / invoice #
        col6 = row.iloc[6] if pd.notna(row.iloc[6]) else None  # Amount

        # Vendor header row
        if col0 and not col0.startswith("Total for") and not col1 and col2 == "":
            # Skip report-level headers
            if col0 not in ("Unpaid Bills Report", "All Dates") and not col0.startswith(
                "Sunday"
            ):
                current_vendor = col0
            continue

        # Detail row (col0 empty, has a transaction type)
        if not col0 and col2 in ("Bill", "Vendor Credit", "Expense") and current_vendor:
            # Only care about Bills for entry purposes
            if col2 == "Bill":
                amount_str = ""
                if col6 is not None:
                    try:
                        amount_str = f"{abs(float(col6)):.2f}"
                    except (ValueError, TypeError):
                        pass

                invoice_num = col3 if col3 not in ("", "nan") else None

                bills.append(
                    {
                        "vendor": current_vendor,
                        "invoice_number": invoice_num,
                        "amount": amount_str,
                    }
                )

    return bills


# ─── Matching ─────────────────────────────────────────────────────────────────


def _build_qb_index(qb_bills: list[dict]) -> tuple[set[str], set[str]]:
    """
    Returns two sets for O(1) lookup:
      vendor_invoice_keys  →  "{norm_vendor}::{norm_invoice}"  (primary)
      vendor_amount_keys   →  "{norm_vendor}::{norm_amount}"   (fallback)
    """
    vi_keys: set[str] = set()
    va_keys: set[str] = set()

    for bill in qb_bills:
        nv = _normalize_vendor(bill["vendor"])
        ni = _normalize(bill["invoice_number"])
        na = _normalize(bill["amount"])

        if nv and ni:
            vi_keys.add(f"{nv}::{ni}")
        if nv and na:
            va_keys.add(f"{nv}::{na}")

    return vi_keys, va_keys


def _is_in_qb(record: PdfNaming, vi_keys: set[str], va_keys: set[str]) -> bool:
    nv = _normalize_vendor(record.vendor)
    ni = _normalize(record.invoice_number)
    na = _normalize(record.amount)

    # Primary match: vendor + invoice number
    if nv and ni:
        if f"{nv}::{ni}" in vi_keys:
            return True

    # Fallback match: vendor + amount
    if nv and na:
        if f"{nv}::{na}" in va_keys:
            return True

    return False


# ─── Route ────────────────────────────────────────────────────────────────────


@router.post("/compare", response_model=CompareResponse)
async def compare_bills(
    file: UploadFile = File(...),
    tenant: Tenant = Depends(require_tenant),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            400, "Please upload a QuickBooks Excel export (.xlsx or .xls)."
        )

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large (10 MB max).")

    # ── Parse the QB export ───────────────────────────────────────────────
    try:
        qb_bills = _parse_qb_excel(file_bytes)
    except Exception as e:
        raise HTTPException(422, f"Could not parse QuickBooks export: {e}")

    if not qb_bills:
        raise HTTPException(
            422,
            "No bills found in the QuickBooks export. Make sure it's an Unpaid Bills Report.",
        )

    vi_keys, va_keys = _build_qb_index(qb_bills)

    # ── Fetch confirmed PDF namings for this tenant ───────────────────────
    result = await db.execute(
        select(PdfNaming)
        .where(
            PdfNaming.tenant_id == tenant.id,
            PdfNaming.confirmed_name.isnot(None),
        )
        .order_by(desc(PdfNaming.created_at))
    )
    records = result.scalars().all()

    # ── Classify each record ──────────────────────────────────────────────
    needs_entry: list[BillItem] = []
    already_entered: list[BillItem] = []
    skipped = 0

    for r in records:
        if r.doc_type in _SKIP_DOC_TYPES:
            skipped += 1
            continue

        item = BillItem(
            confirmed_name=r.confirmed_name or "",
            vendor=r.vendor,
            invoice_number=r.invoice_number,
            amount=r.amount,
            doc_date=r.doc_date,
            doc_type=r.doc_type,
        )

        if _is_in_qb(r, vi_keys, va_keys):
            already_entered.append(item)
        else:
            needs_entry.append(item)

    return CompareResponse(
        needs_entry=needs_entry,
        already_entered=already_entered,
        skipped_receipts=skipped,
        qb_bills_parsed=len(qb_bills),
    )


@router.post("/export.xlsx")
async def export_comparison_excel(
    file: UploadFile = File(...),
    tenant: Tenant = Depends(require_tenant),
    db: AsyncSession = Depends(get_db),
):
    """
    Run the same comparison as /compare and return a styled Excel workbook
    with two sheets: 'Needs Entry' and 'Already in QuickBooks'.
    """
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            400, "Please upload a QuickBooks Excel export (.xlsx or .xls)."
        )

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large (10 MB max).")

    try:
        qb_bills = _parse_qb_excel(file_bytes)
    except Exception as e:
        raise HTTPException(422, f"Could not parse QuickBooks export: {e}")

    if not qb_bills:
        raise HTTPException(422, "No bills found in the QuickBooks export.")

    vi_keys, va_keys = _build_qb_index(qb_bills)

    result = await db.execute(
        select(PdfNaming)
        .where(PdfNaming.tenant_id == tenant.id, PdfNaming.confirmed_name.isnot(None))
        .order_by(desc(PdfNaming.created_at))
    )
    records = result.scalars().all()

    needs_entry: list[BillItem] = []
    already_entered: list[BillItem] = []

    for r in records:
        if r.doc_type in _SKIP_DOC_TYPES:
            continue
        item = BillItem(
            confirmed_name=r.confirmed_name or "",
            vendor=r.vendor,
            invoice_number=r.invoice_number,
            amount=r.amount,
            doc_date=r.doc_date,
            doc_type=r.doc_type,
        )
        if _is_in_qb(r, vi_keys, va_keys):
            already_entered.append(item)
        else:
            needs_entry.append(item)

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Shared styles
    thin = Side(style="thin", color="D1D5DB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(
        ws, bills: list[BillItem], title: str, title_color: str, row_accent: str
    ):
        col_widths = [30, 14, 22, 14, 14, 48]
        headers = ["Vendor", "Bill Date", "Bill No.", "Amount", "Type", "File Name"]

        # Title row
        ws.merge_cells("A1:F1")
        tc = ws["A1"]
        tc.value = f"{title}  —  {len(bills)} bill{'s' if len(bills) != 1 else ''}  ·  Generated {datetime.now().strftime('%m/%d/%Y')}"
        tc.font = Font(bold=True, color="FFFFFF", size=12)
        tc.fill = PatternFill("solid", fgColor=title_color)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # Header row
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor="1E3A5F")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cell_border
        ws.row_dimensions[2].height = 20

        # Column widths
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Data rows
        def _fmt(raw):
            if not raw or not re.match(r"^\d{8}$", raw):
                return raw or ""
            return f"{raw[0:2]}/{raw[2:4]}/{raw[4:8]}"

        total_amount = 0.0
        for i, b in enumerate(bills):
            row = i + 3
            fill = PatternFill("solid", fgColor=row_accent if i % 2 == 0 else "FFFFFF")
            amt = 0.0
            try:
                amt = float(b.amount or 0)
            except ValueError:
                pass
            total_amount += amt

            values = [
                b.vendor or "",
                _fmt(b.doc_date),
                b.invoice_number or "",
                amt,
                b.doc_type or "",
                (b.confirmed_name or "") + ".pdf",
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fill
                c.border = cell_border
                c.font = Font(size=10)
                if col == 4:
                    c.number_format = '"$"#,##0.00'
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif col == 2:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        # Total row
        total_row = len(bills) + 3
        ws.merge_cells(f"A{total_row}:C{total_row}")
        tc = ws.cell(row=total_row, column=1, value="TOTAL")
        tc.font = Font(bold=True, size=10)
        tc.fill = PatternFill("solid", fgColor="E5E7EB")
        tc.border = cell_border
        ac = ws.cell(row=total_row, column=4, value=total_amount)
        ac.number_format = '"$"#,##0.00'
        ac.font = Font(bold=True, size=10)
        ac.fill = PatternFill("solid", fgColor="E5E7EB")
        ac.alignment = Alignment(horizontal="right", vertical="center")
        ac.border = cell_border
        for col in [5, 6]:
            c = ws.cell(row=total_row, column=col)
            c.fill = PatternFill("solid", fgColor="E5E7EB")
            c.border = cell_border

    # Sheet 1: Needs Entry (red/amber theme)
    ws1 = wb.active
    ws1.title = "Needs Entry"
    _write_sheet(ws1, needs_entry, "⚠ Needs Entry in QuickBooks", "B45309", "FEF3C7")

    # Sheet 2: Already Entered (green theme)
    ws2 = wb.create_sheet("Already in QuickBooks")
    _write_sheet(ws2, already_entered, "✓ Already in QuickBooks", "166534", "DCFCE7")

    # Stream response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"QB_Checker_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
